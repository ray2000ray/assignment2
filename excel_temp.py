import sqlite3
import openpyxl
import re
from pylab import *
import matplotlib.pyplot as plt

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = "TemperatureByCity"
# Data can be assigned directly to cells
 #设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
ws.cell("A1").value = "Avgtemp"
ws.cell("B1").value = "City"
ws.cell("C1").value = "Year"
ws.cell("D1").value = "Year_City"
# Rows can also be appended
#ws.append([1, 2, 3])


# Save the file



conn = sqlite3.connect("tempdb.db")
c = conn.cursor()

##what conn here for the sqlite3

##what is need in here
c.execute("SELECT  round(avg(avgTemp),2) as yearAvgTemp, city, strftime('%Y',DATE(date)) year, strftime('%Y',DATE(date))|| city as Year_City FROM GTByCity where avgTemp is not 'None' and country = 'China'group by Year_City order by city,year;")
result = c.fetchall()
rows = len(result)
#print(rows)
start = 0
temps = []
currentCity = result[0][1]
startYear = 1820
#cityNames = []
#cityNames.append(currentCity)
t = arange(0.0, 194, 1)
currentYear = startYear
for row in range(rows):
    print(result[row])
    print("current city:"+currentCity+": current year: "+str(currentYear))
    if currentCity == result[row][1] and currentYear == int(result[row][2]):
        if currentYear < 2013:
            temps.append(result[row][0])
            currentYear = currentYear + 1
        else:
            print(str(currentYear) + "--11: " + str(int(result[row][2])))
            temps.append(result[row][0])
            currentYear = startYear
            print(temps)
            plot(t, temps,label=currentCity)
            plot
#            cityNames.append(currentCity)
            temps = []
            if row < len(result)-1:
                currentCity = result[row+1][1]
            print("set currentcity: "+currentCity + ":" + str(currentYear))
    elif currentCity == result[row][1] and currentYear != int(result[row][2]):
        print(str(currentYear)+"is not " + str(int(result[row][2])))
        for i in range(int(result[row][2])-currentYear+1):
            if currentYear < 2013:
                temps.append(None)
                currentYear = currentYear + 1
                print(str(currentYear) + "--44: " + str(int(result[row][2])))
            else:
                currentYear = currentYear + 1
                temps.append(result[row][0])
                print(str(currentYear) + "--55: " + str(int(result[row][2])))
    elif currentCity != result[row][1] and currentYear != int(result[row][2]):
        print(str(currentYear) + "--is not-- " + str(int(result[row][2])))
        if currentYear < 2013:
            for i in range(int(result[row][2])-currentYear+1):
                temps.append(None)
                currentYear = currentYear + 1
        else:
            print(str(currentYear) + " -22: " + str(int(result[row][2])))
            temps.append(None)
            currentYear = startYear
            temps = []
            currentCity = result[row][1]
    else:
        print("currentYear: "+str(currentYear) + " -22: " +"row year:"+ str(int(result[row][2])))
        print("currentCity: "+currentCity + ": temps len: " + str(len(temps)))
        print(temps)
        plot(t, temps,label=currentCity)
        cityNames.append(currentCity)
        temps = []
        temps.append(result[row][0])
        currentCity = result[row][1]
        currentYear = startYear
        print(currentCity + ":" + str(currentYear))
        #break
    for col in range(4):
        ws.cell(row=row+2, column=col+1).value = result[row][col]


wb.save("WorldTemperature.e.xlsx")

conn.commit()

xlabel('Year')
ylabel('AvgTemperature')
title('AvgTemperature of cities in years')
grid(True)
show()