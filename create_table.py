import sqlite3
import openpyxl

conn = sqlite3.connect("tempdb.db")
c = conn.cursor()

##what conn here for the sqlite3

##what is need in here

c.execute('''CREATE TABLE  stocks(data text, avgTemp text, avgTempUnc text, country text);''')

wb = openpyxl.load_workbook('GlobalLandTemperaturesByCountry.xlsx')

sheet = wb.active
maxrow = sheet.max_row
print('start inserting', end="")
for i in range (2, maxrow+1):
    if (i%10000 == 0):
        print('.', end="", flush = True)
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = str(sheet.cell(row=i, column=2).value)
    avgTempUnc =str(sheet.cell(row=i, column=3).value)
    country=sheet.cell(row=i, column=4).value
    #c.execute("INSERT INTO stocks VALUES ('" + date + "','" + avgTemp + "','"+ avgTempUnc +"','"+ country + "');")
    format_str = """INSERT INTO stocks (data, avgTemp, avgTempUnc, country) 
    	VALUES ("{a1}", "{a2}", "{a3}", "{a4}");"""
    sql_command = format_str.format(a1=date, a2=avgTemp, a3=avgTempUnc, a4=country)
    c.execute(sql_command)
conn.commit()

c.execute("SELECT * FROM stocks WHERE country ='Australia'")
str = c.fetchall()
for s in str:
    print(s)
conn.close()