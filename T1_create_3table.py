import sqlite3
import openpyxl

conn = sqlite3.connect("tempdb.db")
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS GTByCountry(date date, avgTemp DOUBLE, avgTempUnc DOUBLE, country text);''')
c.execute('''CREATE TABLE IF NOT EXISTS GTByCity(date date, avgTemp DOUBLE, avgTempUnc DOUBLE, city text, country text, latitude text, longitude text);''')
c.execute('''CREATE TABLE IF NOT EXISTS GTByState(date date, avgTemp DOUBLE, avgTempUnc DOUBLE, state text, country text);''')

wb = openpyxl.load_workbook('GlobalLandTemperaturesByCountry.xlsx')
sheet = wb.active
maxrow = sheet.max_row
print('start inserting', end="")
for i in range (2, maxrow + 1):
    if (i%10000 == 0):
        print('.', end="", flush = True)
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = str(sheet.cell(row=i, column=2).value)
    avgTempUnc =str(sheet.cell(row=i, column=3).value)
    country=sheet.cell(row=i, column=4).value
    #c.execute("INSERT INTO stocks VALUES ('" + date + "','" + avgTemp + "','"+ avgTempUnc +"','"+ country + "');")
    format_str = """INSERT INTO GTByCountry (date, avgTemp, avgTempUnc, country)
    	VALUES ("{a1}", "{a2}", "{a3}", "{a4}");"""
    sql_command = format_str.format(a1=date, a2=avgTemp, a3=avgTempUnc, a4=country)
    c.execute(sql_command)
conn.commit()

wb = openpyxl.load_workbook('GlobalLandTemperaturesByMajorCity.xlsx')
sheet = wb.active
maxrow = sheet.max_row
print('start inserting', end="")
for i in range (2, maxrow+1):
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = str(sheet.cell(row=i, column=2).value)
    avgTempUnc =str(sheet.cell(row=i, column=3).value)
    city = sheet.cell(row=i, column=4).value
    country = sheet.cell(row=i, column=5).value
    latitude = sheet.cell(row=i, column=6).value
    longitude = sheet.cell(row=i, column=7).value
    #c.execute("INSERT INTO stocks VALUES ('" + date + "','" + avgTemp + "','"+ avgTempUnc +"','"+ country + "');")
    format_str = """INSERT INTO GTByCity (date, avgTemp, avgTempUnc, city, country, latitude, longitude) 
    	VALUES ("{a1}", "{a2}", "{a3}", "{a4}", "{a5}", "{a6}", "{a7}");"""
    sql_command = format_str.format(a1=date, a2=avgTemp, a3=avgTempUnc, a4=city, a5=country, a6=latitude, a7=longitude)
    c.execute(sql_command)
conn.commit()

wb3 = openpyxl.load_workbook('GlobalLandTemperaturesByState.xlsx')
sheet = wb3.active
maxrow = sheet.max_row
print('start inserting', end="")
for i in range (2, maxrow+1):
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = str(sheet.cell(row=i, column=2).value)
    avgTempUnc =str(sheet.cell(row=i, column=3).value)
    state = sheet.cell(row=i, column=4).value
    country = sheet.cell(row=i, column=5).value
    format_str = """INSERT INTO GTByState (date, avgTemp, avgTempUnc, state, country) 
    	VALUES ("{a1}", "{a2}", "{a3}", "{a4}","{a5}");"""
    sql_command = format_str.format(a1=date, a2=avgTemp, a3=avgTempUnc, a4=state, a5=country)
    c.execute(sql_command)
conn.commit()

conn.close()