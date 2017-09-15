# Create a Python script (sql_temp.py) to perform the following tasks:
# List the distinctive major cities located in southern hemisphere ordered by country to the
# console and then write their name, country, and geolocation into a new database table
# called “Southern cities”.
# Find the maximum, minimum and average temperature of Queensland for year 2000 and
# print this information to the console.


import sqlite3
import openpyxl
import re


conn = sqlite3.connect("tempdb.db")
c = conn.cursor()

#create a table to store southern hemisphere cities
c.execute('''CREATE TABLE  Southern_cities(date text, avgTemp text, avgTempUnc text, city text, country text, latitude text, longitude text);''')

wb = openpyxl.load_workbook('GlobalLandTemperaturesByMajorCity.xlsx')
sheet = wb.active
maxrow = sheet.max_row
print('start inserting', end="")
for i in range (2, maxrow+1):
    latitude = sheet.cell(row=i, column=6).value
    #get the cities' latitude  content "S"
    rstr = "S"
    if re.search(rstr, latitude, flags=0):
        date = str(sheet.cell(row=i, column=1).value)
        avgTemp = str(sheet.cell(row=i, column=2).value)
        avgTempUnc =str(sheet.cell(row=i, column=3).value)
        city = sheet.cell(row=i, column=4).value
        country = sheet.cell(row=i, column=5).value
        longitude = sheet.cell(row=i, column=7).value
        format_str = """INSERT INTO Southern_cities (date, avgTemp, avgTempUnc, city, country, latitude, longitude) 
            VALUES ("{a1}", "{a2}", "{a3}", "{a4}", "{a5}", "{a6}", "{a7}");"""
        sql_command = format_str.format(a1=date, a2=avgTemp, a3=avgTempUnc, a4=city, a5=country, a6=latitude, a7=longitude)
        c.execute(sql_command)
conn.commit()

c.execute("SELECT min(avgTemp), max(avgTemp), avg(avgTemp)  FROM GTByCity WHERE state ='Queensland' AND year = AND date is not 'none'")
str = c.fetchall()
for s in str:
    print(s)
conn.close()
