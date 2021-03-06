import sqlite3
import openpyxl
import re
from pylab import *
import matplotlib.pyplot as plt

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = "TemperatureByCity"
# Data can be assigned directly to cells
ws.cell("A1").value = "Avgtemp"
ws.cell("B1").value = "City"
ws.cell("C1").value = "Year"
#ws.cell("D1").value = "Year_City"
# Rows can also be appended
#ws.append([1, 2, 3])


# Save the file



conn = sqlite3.connect("tempdb.db")
c = conn.cursor()


c.execute("SELECT  round(avg(avgTemp),2) as yearAvgTemp, state, strftime('%Y',DATE(date)) year, \
strftime('%Y',DATE(date))|| state as Year_state  \
FROM GTByState where avgTemp is not 'None' and country = 'China' group by Year_state order by state,year;")
result = c.fetchall()
rows = len(result)
#print(rows)
#start = 0
temps = []
currentCity = result[0][1]
startYear = 1820
cityNames = []


## method 1 to implement line graph function by using plot

t = arange(0.0, 194, 1)
currentYear = startYear
for row in range(rows):
    #print(result[row])
    #print("current city:"+currentCity+": current year: "+str(currentYear))
    if currentCity == result[row][1] and currentYear == int(result[row][2]):
        if currentYear < 2013:
            temps.append(result[row][0])
            currentYear = currentYear + 1
        else:
            print(str(currentYear) + "--finish--: " + str(int(result[row][2])))
            print("current city:" + currentCity + ": current year: " + str(currentYear))
            temps.append(result[row][0])
            cityNames.append(currentCity)
            currentYear = startYear
            print(temps)
            #using different line styles for indicating different cities
            if row <1600:
                plt.plot(t+1820, temps, linestyle="-", label=currentCity)
            elif row< 3200:
                plt.plot(t + 1820, temps, linestyle="--", label=currentCity)
            else:
                plt.plot(t + 1820, temps, linestyle=":", label=currentCity)
#            cityNames.append(currentCity)
            temps = []
            if row < len(result)-1:
                currentCity = result[row+1][1]
            #print("set currentcity: "+currentCity + ":" + str(currentYear))
    elif currentCity == result[row][1] and currentYear != int(result[row][2]):
        #print(str(currentYear)+"is not " + str(int(result[row][2])))
        for i in range(int(result[row][2])-currentYear+1):
            if currentYear < 2013:
                temps.append(None)

                currentYear = currentYear + 1
                #print(str(currentYear) + "--44: " + str(int(result[row][2])))
            else:
                currentYear = currentYear + 1
                temps.append(result[row][0])
                #print(str(currentYear) + "--55: " + str(int(result[row][2])))
    elif currentCity != result[row][1] and currentYear != int(result[row][2]):
        print(str(currentYear) + "--is not-- " + str(int(result[row][2])))
        if currentYear < 2013:
            for i in range(int(result[row][2])-currentYear+1):
                temps.append(None)
                currentYear = currentYear + 1
        else:
            print(str(currentYear) + " -22: " + str(int(result[row][2])))
            temps.append(None)
            currentYear = startYear
            temps = []
            currentCity = result[row][1]
    else: # this section used to comlepted the if logical
        #print("currentYear: "+str(currentYear) + " -22: " +"row year:"+ str(int(result[row][2])))
        # #print("currentCity: "+currentCity + ": temps len: " + str(len(temps)))
        # print(str(currentYear) + "--finish--: " + str(int(result[row][2])))
        # print("current city:" + currentCity + ": current year: " + str(currentYear))
        # print(temps)
        cityNames.append(currentCity)
        plt.plot(t + 1820, temps, linestyle="-", label=currentCity)
        temps = []
        temps.append(result[row][0])
        currentCity = result[row][1]
        currentYear = startYear
        #print(currentCity + ":" + str(currentYear))
    for col in range(3):
        ws.cell(row=row+2, column=col+1).value = result[row][col]


print(cityNames)



wb.save("WorldTemperature.xlsx")

conn.commit()

xlabel('Year')
ylabel('AvgTemperature')
title('AvgTemperature of cities of China in years')
grid(True)
plt.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
plt.show()