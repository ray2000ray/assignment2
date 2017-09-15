import sqlite3
import openpyxl
import matplotlib.pyplot as plt
from pylab import *

wb = openpyxl.load_workbook('WorldTemperature.xlsx')
wb.create_sheet("Comparison")

ws = wb.get_sheet_by_name('Comparison')
ws.cell("A1").value = "Difference"
ws.cell("B1").value = "State"
ws.cell("C1").value = "Year"
# ws["A1", "B1", "C1"] = ["Difference", "Year", "State"]

conn = sqlite3.connect("tempdb.db")
c = conn.cursor()
#part one

c.execute("select round(s.yearAvgTemp -  c.yearAvgTemp,2)  as difference,  s.state, s.year \
 from (SELECT round(avg(avgTemp),2) as yearAvgTemp, state, strftime('%Y',DATE(date)) year, strftime('%Y',DATE(date))|| state as Year_state \
FROM GTByState \
where avgTemp is not 'None' and country = 'Australia' \
group by Year_state order by state, year) as s,  (SELECT  round(avg(avgTemp),2) as yearAvgTemp, country, strftime('%Y',DATE(date)) year \
FROM GTByCountry \
where avgTemp is not 'None' and country = 'Australia' \
group by year \
order by year) as c \
where s.year =  c.year \
order by s.state, s.year;")


result = c.fetchall()
rows = len(result)

#
# for row in range(rows):
#
#     for col in range(3):
#         ws.cell(row=row+2, column=col+1).value = result[row][col]
#####

temps =[]
stateNames = []
currentState = result[0][1]
startYear = 1952
t = arange(0.0, 62, 1)
currentYear = startYear
for row in range(rows):
    #print(result[row])
    #print("current city:"+currentState+": current year: "+str(currentYear))
    if currentState == result[row][1] and currentYear == int(result[row][2]):
        if currentYear < 2013:
            temps.append(result[row][0])
            currentYear = currentYear + 1
        else:
            #print(str(currentYear) + "--11: " + str(int(result[row][2])))
            temps.append(result[row][0])
            currentYear = startYear
            print(temps)
            stateNames.append(currentState)
            plot(t+1952, temps, linestyle="-", label=currentState)
#            cityNames.append(currentState)
            temps = []
            if row < len(result)-1:
                currentState = result[row+1][1]
            #print("set currentState: "+currentState + ":" + str(currentYear))
    elif currentState == result[row][1] and currentYear != int(result[row][2]):
        #print(str(currentYear)+"is not " + str(int(result[row][2])))
        for i in range(int(result[row][2])-currentYear+1):
            if currentYear < 2013:
                temps.append(None)

                currentYear = currentYear + 1
                #print(str(currentYear) + "--44: " + str(int(result[row][2])))
            else:
                currentYear = currentYear + 1
                temps.append(result[row][0])
                #print(str(currentYear) + "--55: " + str(int(result[row][2])))
    elif currentState != result[row][1] and currentYear != int(result[row][2]):
        print(str(currentYear) + "--is not-- " + str(int(result[row][2])))
        if currentYear < 2013:
            for i in range(int(result[row][2])-currentYear+1):
                temps.append(None)
                currentYear = currentYear + 1
        else:
            print(str(currentYear) + " -22: " + str(int(result[row][2])))
            temps.append(None)
            currentYear = startYear
            temps = []
            currentState = result[row][1]
    else:
        #print("currentYear: "+str(currentYear) + " -22: " +"row year:"+ str(int(result[row][2])))
        #print("currentState: "+currentState + ": temps len: " + str(len(temps)))
        print(temps)
        stateNames.append(currentState)
        plt.plot(t + 1952, temps, linestyle="-", label=currentState)
        temps = []
        temps.append(result[row][0])
        currentState = result[row][1]
        currentYear = startYear
        #print(currentState + ":" + str(currentYear))
    for col in range(3):
        ws.cell(row=row+2, column=col+1).value = result[row][col]

wb.save("WorldTemperature.xlsx")

xlabel('Year')
ylabel('Difference')
title('Comparison Difference temp of states in Australia by years')
grid(True)
show()